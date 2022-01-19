import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


def DfReadDataFile(filename):
    """
    Parameters
    ----------
    filename : JAFROC format Excel input file

    Returns
    -------
    dataset object

    """
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames
    expectedNames=["NL","LL","TRUTH"]
    msg = "Excel workbook is missing at least one of NL, LL or "
    "TRUTH worksheets."
    myExit(sheetNames,expectedNames,msg)

    ws = wb['TRUTH']
    data = ws.values

    columnNames = next(data)[0:]
    expectedNames = ['CaseID', 'LesionID', 'Weight']
    msg = ("Excel workbook TRUTH sheet has missing or incorrect "
        "required column names. These are the correct names: "
        " 'CaseID', 'LesionID', 'Weight'")
    myExit(columnNames,expectedNames,msg)

    df = pd.DataFrame(data, columns=columnNames)
    
    if df.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")
        
    df["TruthID"]=(df["LesionID"] > 0).astype(int)
    df = df.sort_values(["TruthID", "CaseID"])
    caseIDCol = df["CaseID"]
    lesionIDCol = df["LesionID"]
    weightCol = df["Weight"]
    L=len(caseIDCol)
    allCases = np.unique(np.array(df["CaseID"]))
    normalCases = df.loc[df['LesionID'] == 0]["CaseID"]
    K1=len(normalCases)
    abnormalCases = df.loc[df['LesionID'] == 1]["CaseID"]
    K2=len(abnormalCases)
    K = K1 + K2
    
    # calculate lesion perCase
    x=pd.Series(df["CaseID"])
    x=x.isin(abnormalCases)
    x=pd.Series(df["CaseID"][x])
    x=x.value_counts()
    perCase=x.sort_index()
    
    # test code with unsorted file, where CaseID is not neatly ordered
    
    
    return(df)
