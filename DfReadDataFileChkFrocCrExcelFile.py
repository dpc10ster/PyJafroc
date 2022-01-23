import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


def testDfReadDataFile(filename):
    DfReadDataFile(filename)


def DfCheck(filename):

# =============================================================================
# Tested with 1 good and 2 bad files
# filename = 'extdata/toyFiles/FROC/frocCr.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx' unordered TRUTH
# filename = 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx' unordered TRUTH
# filename = 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx' incorrect sheet names
# filename = 'extdata/toyFiles/FROC/bad/frocCr-04.xlsx' normal case in LL
# filename = 'extdata/toyFiles/FROC/bad/frocCr-05.xlsx' do: numeric format
# filename = 'extdata/toyFiles/FROC/bad/frocCr2BlankRows.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCrNonCharInReaderID.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/incorrectCaseIDsInLL.xlsx' why missing?
# filename = 'extdata/toyFiles/FROC/bad/incorrectCaseIDsInLL2.xlsx'
# filename = "extdata/toyFiles/FROC/bad/incoCaseIDsInTP.xlsx"
# =============================================================================

    wb = load_workbook(filename)
    sheetNames = wb.sheetnames
    expectedNames = ["NL", "LL", "TRUTH"]
    msg = ("Excel workbook has missing or incorrectly named sheets. "
           "These are the correct names: ") + ", ".join(expectedNames)
    myExit(sheetNames, expectedNames, msg)

    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]

    expectedNames = ['CaseID', 'LesionID', 'Weight']
    msg = ("Excel worksheet TRUTH has missing or incorrect "
           "required column names. "
           "These are the correct names: ") + ", ".join(expectedNames)
    myExit(columnNames, expectedNames, msg)

    dfTruth = pd.DataFrame(data, columns=columnNames)
    # Check for missing cells
    if dfTruth.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")

    AllCases = np.unique(dfTruth["CaseID"])
    NormalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(NormalCases)
    AbnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(AbnormalCases)
    K = K1 + K2
    
    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

    expectedNames = ['ReaderID', 'ModalityID', 'CaseID', 'NLRating']
    msg = ("Excel worksheet NL has missing or incorrect "
           "required column names. "
           "These are the correct names: ") + ", ".join(expectedNames)
        
    if dfNL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in NL worksheet")

    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)
    

    if dfLL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in LL worksheet")

    # check for occurence of normal cases in LL sheet
    x1 = dfLL["CaseID"]
    x1 = set(x1.astype(int))
    x2 = NormalCases
    x2 = set(x2.astype(int))
    if len(x1 & x2) != 0:
        sys.exit("Normal cases encountered in LL worksheet")
        
# =============================================================================
# TODO: Add checks for duplicate rows in LL sheet
# =============================================================================

def DfReadDataFile(filename):
    """
    Parameters
    ----------
    filename : JAFROC format Excel input file

    Returns
    -------
    dataset object

    """
# =============================================================================
# Check the Excel file
# Load the Excel file
# =============================================================================
    DfCheck(filename)
    wb = load_workbook(filename)

# =============================================================================
# Load the TRUTH worksheet
# Exract the columns
# =============================================================================
    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfTruth = pd.DataFrame(data, columns=columnNames)
    # add TruthID column based on cases with 0 or > 0 lesions
    dfTruth["TruthID"] = (dfTruth["LesionID"] > 0).astype(int)


# =============================================================================
# dfTruth
# Out[732]: 
#     CaseID  LesionID Weight  TruthID
# 0        1         0      0        0
# 1        2         0      0        0
# 2        3         0      0        0
# 3       70         1    0.3        1
# 4       70         2    0.7        1
# 5       71         1      1        1
# 6       72         1  0.333        1
# 7       72         2  0.333        1
# 8       72         3  0.333        1
# 9       73         1    0.1        1
# 10      73         2    0.9        1
# 11      74         1      1        1
# =============================================================================
    
    # sort on "TruthID" & "CaseID" fields to put non-diseased cases first
    dfTruth = dfTruth.sort_values(["TruthID", "CaseID"])
    # TODO variable weights not currently implemented
    dfTruth['Weight'] = dfTruth['Weight'].astype(float, errors = 'raise')
    # weightCol = dfTruth["Weight"]

# =============================================================================
# dfTruth
# Out[740]: 
#     CaseID  LesionID  Weight  TruthID
# 0        1         0   0.000        0
# 1        2         0   0.000        0
# 2        3         0   0.000        0
# 3       70         1   0.300        1
# 4       70         2   0.700        1
# 5       71         1   1.000        1
# 6       72         1   0.333        1
# 7       72         2   0.333        1
# 8       72         3   0.333        1
# 9       73         1   0.100        1
# 10      73         2   0.900        1
# 11      74         1   1.000        1
# =============================================================================
        
    AllCases = np.unique(dfTruth["CaseID"])
    NormalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(NormalCases)
    AbnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(AbnormalCases)
    K = K1 + K2
    
    # calculate lesion perCase
    x = pd.Series(dfTruth["CaseID"])
    x = x.isin(AbnormalCases)
    x = pd.Series(dfTruth["CaseID"][x])
    x = x.value_counts()
    perCase = x.sort_index()

    maxLL = max(perCase)
    relWeights = [1/maxLL] * maxLL
 
# =============================================================================
# Load the NL sheet
# Exract the columns
# =============================================================================
    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

# =============================================================================
# Load the LL sheet
# Exract the columns
# =============================================================================
    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)
    
    # with FROC data not all modalities may appear in NL and LL sheets
    modalities = (dfLL["ModalityID"].append(dfNL["ModalityID"])).unique()
    I = len(modalities)
    readers = (dfLL["ReaderID"].append(dfNL["ReaderID"])).unique()
    J = len(readers)
    lesions = np.unique(dfTruth["LesionID"])[1:]
    maxNL = dfNL.groupby(['ReaderID',
                          'ModalityID', 
                          'CaseID']).transform(len).max()[0]
    
    maxLL = max(perCase)
    
    # construct truthTabiLeStr
    truthTableStr = np.full((I,J,K,maxLL+1), 0)
    for indxCsId in range(len(dfTruth["CaseID"])):
        c = (AllCases == dfTruth["CaseID"][indxCsId])
        l = dfTruth["LesionID"][indxCsId]
        truthTableStr[:, :, c, l] = 1
    
# =============================================================================
#         
#     (truthTableStr[:,:,0:2,0]).all()      True
#     (truthTableStr[:,:,0:2,1:]).all()     False   
#     (truthTableStr[:,:,3:,0]).all()       False
#     (truthTableStr[:,:,3:,1]).all()       True
#     (truthTableStr[:,:,[3,5,7],1]).all()  True
#     (truthTableStr[:,:,[3,5,6],2]).all()  True
#     (truthTableStr[:,:,5,3]).all()        True
# 
#     
# =============================================================================
    NL = np.full((I,J,K,maxNL), -np.inf)
    for indxNl in range(len(dfNL["ModalityID"])):
        i = (modalities == dfNL["ModalityID"][indxNl])
        j = (readers == dfNL["ReaderID"][indxNl])
        c = (AllCases == dfNL["CaseID"][indxNl])

# =============================================================================
#         if dfNL["CaseID"][indxNl] in NormalCases:
#             tt2 = truthTableStr[i,j,c,1] 
#         else: 
#             tt2 = truthTableStr[i,j,c,2]   
# =============================================================================
        matchCount = ((dfNL["CaseID"] == dfNL["CaseID"][indxNl]) & 
                      (dfNL["ModalityID"] == dfNL["ModalityID"][indxNl]) & 
                      (dfNL["ReaderID"] == dfNL["ReaderID"][indxNl]))
        for l in range(sum(matchCount)):
            if NL[i, j, c, l] == -np.inf: 
                NL[i, j, c, l] = dfNL["NLRating"][indxNl + l]
            truthTableStr[i, j, c, l] = 1


# =============================================================================
# NL[0, 0, :, :]
# Out[801]: 
# array([[1.02, 2.17],
#        [2.22, -inf],
#        [1.9 , -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf]])
# 
# NL[0, 1, :, :]
# Out[802]: 
# array([[2.21, -inf],
#        [3.1 , 2.21],
#        [2.07, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf]])
# 
# NL[0, 2, :, :]
# Out[803]: 
# array([[2.14, -inf],
#        [1.98, -inf],
#        [1.95, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf]])
# 
# NL[1, 0, :, :]
# Out[804]: 
# array([[2.89, -inf],
#        [2.89, -inf],
#        [3.22, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [1.85, -inf],
#        [0.84, -inf]])
# 
# NL[1, 1, :, :]
# Out[805]: 
# array([[3.01, -inf],
#        [1.96, -inf],
#        [2.08, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf]])
# 
# NL[1, 2, :, :]
# Out[806]: 
# array([[-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [-inf, -inf],
#        [2.24, 4.01],
#        [1.86, -inf],
#        [-inf, -inf],
#        [-inf, -inf]])
# 
#             
# =============================================================================
    
    LL = np.full((I,J,K2,maxLL), -np.inf)
    for indxLl in range(len(dfLL["ModalityID"])):
        i = (modalities == dfLL["ModalityID"][indxLl])
        j = (readers == dfLL["ReaderID"][indxLl])
        c = (AbnormalCases == dfLL["CaseID"][indxLl])
        #l = (lesions == dfLL["LesionID"][indxLl])
        
# =============================================================================
#         if K1 != 0:
#             l = np.iL(np.unique(dfTruth["LesionID"])[1:] == 
#                             dfTruth["LesionID"][l])
#         else:    
#             l = np.iL(np.unique(dfTruth["LesionID"]) == 
#                             dfTruth["LesionID"][l])
# 
# =============================================================================
        matchCount = ((dfLL["CaseID"] == dfLL["CaseID"][indxLl]) & 
                      (dfLL["ModalityID"] == dfLL["ModalityID"][indxLl]) & 
                      (dfLL["ReaderID"] == dfLL["ReaderID"][indxLl]))
        for l in range(sum(matchCount)):
            if LL[i, j, c, l] == -np.inf:
                LL[i, j, c, l] = dfLL["LLRating"][indxLl + l]
            truthTableStr[i, j, np.append([False]*3, c), l] = 1
    
    
# =============================================================================
#   LL[0, 0, :, :]
#   Out[834]: 
#   array([[5.28, 4.65, -inf],
#          [3.01, -inf, -inf],
#          [5.98, -inf, -inf],
#          [5.  , 5.25, -inf],
#          [4.26, -inf, -inf]])
# 
#   LL[0, 1, :, :]
#   Out[835]: 
#   array([[5.14, -inf, -inf],
#          [3.31, -inf, -inf],
#          [4.92, 5.11, 4.63],
#          [4.95, -inf, -inf],
#          [5.3 , -inf, -inf]])
# 
#   LL[0, 2, :, :]
#   Out[836]: 
#   array([[4.66, -inf, -inf],
#          [4.03, -inf, -inf],
#          [5.22, -inf, -inf],
#          [4.94, -inf, -inf],
#          [5.27, -inf, -inf]])
# 
#   LL[1, 0, :, :]
#   Out[837]: 
#   array([[5.2 , -inf, -inf],
#          [3.27, -inf, -inf],
#          [4.61, -inf, -inf],
#          [5.18, -inf, -inf],
#          [4.72, -inf, -inf]])
# 
#   LL[1, 1, :, :]
#   Out[838]: 
#   array([[4.77, -inf, -inf],
#          [3.19, -inf, -inf],
#          [5.2 , -inf, -inf],
#          [5.39, -inf, -inf],
#          [5.01, -inf, -inf]])
# 
#   LL[1, 2, :, :]
#   Out[839]: 
#   array([[4.87, -inf, -inf],
#          [1.94, -inf, -inf],
#          [-inf, -inf, -inf],
#          [-inf, -inf, -inf],
#          [-inf, -inf, -inf]])
#     
#   
# =============================================================================
    return(dfTruth)


