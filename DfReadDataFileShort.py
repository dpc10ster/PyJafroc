import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.simplefilter("ignore")



# =============================================================================
# Tested with 1 good and 2 bad files
# filename = 'extdata/toyFiles/FROC/frocCr.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx' unexpected case
# filename = 'extdata/toyFiles/FROC/bad/frocCr-04.xlsx' normal case in LL
# fn = ['extdata/toyFiles/FROC/frocCr.xlsx',
# 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx',
# 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx',
# 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx',
# 'extdata/toyFiles/FROC/bad/frocCr-04.xlsx']
# =============================================================================


def DfReadDataFile(filename):
    """
    Parameters
    ----------
    filename : JAFROC format Excel input file

    Returns
    -------
    dataset object

    """
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames

    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]

    # Extract the data minus the column names
    dfTruth = pd.DataFrame(data, columns=columnNames)

    dfTruth["TruthID"] = (dfTruth["LesionID"] > 0).astype(int)
    # sort on "TruthID" & "CaseID" fields, putting non-diseased cases first
    dfTruth = dfTruth.sort_values(["TruthID", "CaseID"])
    truthCaseIDCol = dfTruth["CaseID"]
    truthLesionIDCol = dfTruth["LesionID"]
    weightCol = dfTruth["Weight"]
    L = len(truthCaseIDCol)
    allCases = np.unique(np.array(dfTruth["CaseID"]))
    normalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(normalCases)
    abnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(abnormalCases)
    K = K1 + K2

    # calculate lesion perCase
    x = pd.Series(dfTruth["CaseID"])
    x = x.isin(abnormalCases)
    x = pd.Series(dfTruth["CaseID"][x])
    x = x.value_counts()
    perCase = x.sort_index()

    maxLL = max(perCase)
    relWeights = [1/maxLL] * maxLL

# =============================================================================
# Read LL sheet and check that all column names exits
# =============================================================================
    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]


    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)
    
    # normal cases cannot occur in LL sheet
    if bool(set(dfLL["CaseID"]) & set(normalCases)):
        sys.exit("normal cases cannot occur in LL sheet")

    # Check for missing cells
    if dfNL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in LL worksheet")

    llReaderIDCol = dfLL["ReaderID"]
    llModalityIDCol = dfLL["ModalityID"]
    llCaseIDCol = dfLL["CaseID"]
    llLesionIDCol = dfLL["LesionID"]
    llRatingsCol = dfLL["LLRating"]

    return(dfTruth)


def testDfReadDataFile(filename):
    DfReadDataFile(filename)
