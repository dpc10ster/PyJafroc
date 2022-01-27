import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


# =============================================================================
# Put all checks of the Excel file in here
# =============================================================================
def DfCheck(FileName):

# =============================================================================
# Tested with 1 good and 2 bad files
# FileName = 'extdata/toyFiles/FROC/frocCr.xlsx'
# FileName = 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx' unordered TRUTH
# FileName = 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx' unordered TRUTH
# FileName = 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx' incorrect sheet names
# FileName = 'extdata/toyFiles/FROC/bad/frocCr-04.xlsx' normal case in LL
# FileName = 'extdata/toyFiles/FROC/bad/frocCr-05.xlsx' do: numeric format
# FileName = 'extdata/toyFiles/FROC/bad/frocCr2BlankRows.xlsx'
# FileName = 'extdata/toyFiles/FROC/bad/frocCrNonCharInReaderID.xlsx'
# FileName = 'extdata/toyFiles/FROC/bad/incorrectCaseIDsInLL.xlsx' why missing?
# FileName = 'extdata/toyFiles/FROC/bad/incorrectCaseIDsInLL2.xlsx'
# FileName = "extdata/toyFiles/FROC/bad/incoCaseIDsInTP.xlsx"
# =============================================================================

    wb = load_workbook(FileName)
    sheetNames = wb.sheetnames
    expectedNames = ["NL", "LL", "TRUTH"]
    msg = ("Excel workbook has missing or incorrectly named sheets. "
           "These are the correct names: ") + ", ".join(expectedNames)
    myExit(sheetNames, expectedNames, msg)

    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]

    expectedNames = ['CaseID', 'LesionID', 'Weight']
    msg = ("Excel worksheet TRUTH has missing or incorrect "
           "required column names. "
           "These are the correct names: ") + ", ".join(expectedNames)
    myExit(columnNames, expectedNames, msg)

    dfTruth = pd.DataFrame(data, columns=columnNames)
    # Check for missing cells
    if dfTruth.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")

    AllCases = np.unique(dfTruth["CaseID"])
    NormalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(NormalCases)
    AbnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(AbnormalCases)
 
    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

    expectedNames = ['ReaderID', 'ModalityID', 'CaseID', 'NLRating']
    msg = ("Excel worksheet NL has missing or incorrect "
           "required column names. "
           "These are the correct names: ") + ", ".join(expectedNames)

    if dfNL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in NL worksheet")

    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)

    if dfLL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in LL worksheet")

    # check for occurence of normal cases in LL sheet
    x1 = dfLL["CaseID"]
    x1 = set(x1.astype(int))
    x2 = NormalCases
    x2 = set(x2.astype(int))
    if len(x1 & x2) != 0:
        sys.exit("Normal cases encountered in LL worksheet")


# =============================================================================
# TODO: Add checks for duplicate rows in LL sheet
# =============================================================================

def DfReadDataFile(FileName, DataType="FROC"):
    """
    Parameters
    ----------
    FileName : str
        JAFROC format Excel input file name

    DataType : str
        The type of data, "FROC" (default) or "ROC"

    Returns
    -------
    dataset object ds

    """
# =============================================================================
# Check the Excel file
# Load the Excel file
# =============================================================================
    DfCheck(FileName)
    wb = load_workbook(FileName)

# =============================================================================
# Load the TRUTH worksheet
# Extract the columns
# =============================================================================
    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfTruth = pd.DataFrame(data, columns=columnNames).astype(int)
    # add TruthID column based on cases with 0 or > 0 lesions
    dfTruth["TruthID"] = (dfTruth["LesionID"] > 0).astype(int)


# =============================================================================
# See DfReadDataFileChkFrocCrExcelFile.py for cross check with input file
# 'extdata/toyFiles/FROC/frocCr.xlsx'
# =============================================================================

    # sort on "TruthID" & "CaseID" fields to put non-diseased cases first
    dfTruth = dfTruth.sort_values(["TruthID", "CaseID"])
    # TODO variable weights not currently implemented
    # dfTruth['Weight'] = dfTruth['Weight'].astype(float, errors='raise')
    # weightCol = dfTruth["Weight"]

    u, ind = np.unique(dfTruth["CaseID"], return_index=True)
    AllCases = u[np.argsort(ind)]
    NormalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(NormalCases)
    AbnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(AbnormalCases)
    K = K1 + K2
    
    # calculate lesion perCase
    x = pd.Series(dfTruth["CaseID"])
    x = x.isin(AbnormalCases)
    x = pd.Series(dfTruth["CaseID"][x])
    x = x.value_counts(sort = False)
    perCase = x.sort_index()
    # next line indexes using abnormal cases, 0 to K2-1
    # Fixes indexing of perCase array
    perCase = pd.Series(list(perCase))

    maxLL = max(perCase)
    relWeights = [1/maxLL] * maxLL

# =============================================================================
# Load the NL sheet
# Extract the columns
# =============================================================================
    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

# =============================================================================
# Load the LL sheet
# Exract the columns
# =============================================================================
    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)

    # with FROC data not all modalities may appear in NL and LL sheets
    modalities = (dfLL["ModalityID"].append(dfNL["ModalityID"])).unique()
    I = len(modalities)
    readers = (dfLL["ReaderID"].append(dfNL["ReaderID"])).unique()
    J = len(readers)
    # lesions = np.unique(dfTruth["LesionID"])[1:]
    maxNL = dfNL.groupby(['ReaderID',
                          'ModalityID',
                          'CaseID']).transform(len).max()[0]

    maxLL = max(perCase)

    NL = np.full((I,J,K,maxNL), -np.inf)
    for indxNl in range(len(dfNL["ModalityID"])):
        i = (modalities == dfNL["ModalityID"][indxNl])
        j = (readers == dfNL["ReaderID"][indxNl])
        c = (AllCases == dfNL["CaseID"][indxNl])

        matchCount = ((dfNL["CaseID"] == dfNL["CaseID"][indxNl]) & 
                      (dfNL["ModalityID"] == dfNL["ModalityID"][indxNl]) & 
                      (dfNL["ReaderID"] == dfNL["ReaderID"][indxNl]))
        for l in range(sum(matchCount)):
            if NL[i, j, c, l] == -np.inf: 
                NL[i, j, c, l] = dfNL["NLRating"][indxNl + l]

    lesions = list(range(1, maxLL+1))
    LL = np.full((I,J,K2,maxLL), -np.inf)
    for indxLl in range(len(dfLL["ModalityID"])):
        i = (modalities == dfLL["ModalityID"][indxLl])
        j = (readers == dfLL["ReaderID"][indxLl])
        c = (AbnormalCases == dfLL["CaseID"][indxLl])
        l = (lesions == dfLL["LesionID"][indxLl])
        
        LL[i, j, c, l] = dfLL["LLRating"][indxLl]

# =============================================================================
# Return a dataset object
# =============================================================================
    ds = [NL, LL, perCase, relWeights, DataType]
    return(ds)


# FileName = "extdata/toyFiles/FROC/frocCr.xlsx"
# FileName = "extdata/JT.xlsx"
# ds = DfReadDataFile(FileName)
