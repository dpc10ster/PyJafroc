import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


# =============================================================================
# Tested with 1 good and 2 bad files
# filename = 'extdata/toyFiles/FROCfrocCr.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx' unexpected case 
# =============================================================================


def DfReadDataFile(filename):
    """
    Parameters
    ----------
    filename : JAFROC format Excel input file

    Returns
    -------
    dataset object

    """
# =============================================================================
# Load the Excel sheet
# and check that all required worksheets exist
# =============================================================================
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames
    sheetNamesU = [each_string.upper() for each_string in sheetNames]
    expectedNames = ["NL", "LL", "TRUTH"]
    msg = "Excel workbook is missing at least one of NL, LL or "
    "TRUTH worksheets."
    myExit(sheetNamesU, expectedNames, msg)

# =============================================================================
# Load the TRUTH worksheet
# and check that all required columnNames exist
# =============================================================================
    ws = wb['TRUTH']
    data = ws.values

    columnNames = next(data)[0:]
    columnNamesU = [each_string.upper() for each_string in columnNames]
    expectedNames = ['CaseID', 'LesionID', 'Weight']
    expectedNamesU = [each_string.upper() for each_string in expectedNames]
    msg = ("Excel workbook TRUTH sheet has missing or incorrect "
           "required column names. These are the correct names: "
           " 'CaseID', 'LesionID', 'Weight'")
    myExit(columnNamesU, expectedNamesU, msg)

    # Extract the data minus the column names
    df = pd.DataFrame(data, columns=columnNames)

    # Check for missing cells
    if df.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")

    df["TruthID"] = (df["LesionID"] > 0).astype(int)
    # sort on "TruthID" & "CaseID" fields, putting non-diseased cases first
    df = df.sort_values(["TruthID", "CaseID"])
    caseIDCol = df["CaseID"]
    lesionIDCol = df["LesionID"]
    weightCol = df["Weight"]
    L = len(caseIDCol)
    allCases = np.unique(np.array(df["CaseID"]))
    normalCases = df.loc[df['LesionID'] == 0]["CaseID"]
    K1 = len(normalCases)
    abnormalCases = df.loc[df['LesionID'] == 1]["CaseID"]
    K2 = len(abnormalCases)
    K = K1 + K2

    # calculate lesion perCase
    x = pd.Series(df["CaseID"])
    x = x.isin(abnormalCases)
    x = pd.Series(df["CaseID"][x])
    x = x.value_counts()
    perCase = x.sort_index()

    maxLL = max(perCase)
    relWeights = [1/maxLL] * maxLL
 
    
 
    return(df)
