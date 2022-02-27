import warnings
import sys
import pandas as pd
import numpy as np
import copy
from openpyxl import load_workbook
warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


# =============================================================================
# Put all checks of the Excel file in here
# =============================================================================
def DfCheckDataFile(FileName):

    wb = load_workbook(FileName)
    sheetNames = wb.sheetnames
    expectedNames = ["NL", "LL", "TRUTH"]
    msg = ("Excel workbook has missing or incorrectly named sheets. "
           "The expected names (case sensitive) are: ") + ", ".join(expectedNames)
    myExit(sheetNames, expectedNames, msg)

    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]

    expectedNames = ['CaseID', 'LesionID', 'Weight']
    msg = ("Excel worksheet TRUTH has missing or incorrect "
           "column names. "
           "The expected names are: ") + ", ".join(expectedNames)
    myExit(columnNames, expectedNames, msg)

    dfTruth = pd.DataFrame(data, columns=columnNames)
    # Check for missing cells
    if dfTruth.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")

    AllCases = np.unique(dfTruth["CaseID"])
    NormalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(NormalCases)
    AbnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(AbnormalCases)

    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

    expectedNames = ['ReaderID', 'ModalityID', 'CaseID', 'NLRating']
    msg = ("Excel worksheet NL has missing or incorrect "
           "required column names. "
           "These are the correct names: ") + ", ".join(expectedNames)

    if dfNL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in NL worksheet")

    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)

    if dfLL.isnull().values.any():
        sys.exit("Missing cell(s) encountered in LL worksheet")

    # check for occurence of normal cases in LL sheet
    x1 = dfLL["CaseID"]
    x1 = set(x1.astype(int))
    x2 = NormalCases
    x2 = set(x2.astype(int))
    if len(x1 & x2) != 0:
        sys.exit("Normal cases encountered in LL worksheet")


# =============================================================================
# TODO: Add checks for duplicate rows in LL sheet
# =============================================================================

def DfReadDataFile(FileName, DataType="FROC"):
    """
    Parameters
    ----------
    FileName : str
        JAFROC format Excel input file name

    DataType : str
        The type of data, "FROC" (default) or "ROC"

    Returns
    -------
    dataset list object ds = [NL, LL, perCase, relWeights, DataType]

    """
# =============================================================================
# Check the Excel file
# Load the Excel file
# =============================================================================
    DfCheckDataFile(FileName)
    wb = load_workbook(FileName)

# =============================================================================
# Load the TRUTH worksheet
# Extract the columns
# =============================================================================
    ws = wb['TRUTH']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfTruth = pd.DataFrame(data, columns=columnNames).astype(int)
    # add TruthID column based on cases with 0 or > 0 lesions
    dfTruth["TruthID"] = (dfTruth["LesionID"] > 0).astype(int)


# =============================================================================
# See DfReadDataFileChkFrocCrExcelFile.py for cross check with input file
# 'extdata/toyFiles/FROC/frocCr.xlsx'; 
# in /Users/Dev/Desktop/PyJafroc scraps/moved
# =============================================================================

    # sort on "TruthID" & "CaseID" fields to put non-diseased cases first
    dfTruth = dfTruth.sort_values(["TruthID", "CaseID"])
    # TODO variable weights not currently implemented
    # dfTruth['Weight'] = dfTruth['Weight'].astype(float, errors='raise')
    # weightCol = dfTruth["Weight"]

    u, ind = np.unique(dfTruth["CaseID"], return_index=True)
    AllCases = u[np.argsort(ind)]
    NormalCases = np.array(dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"])
    K1 = len(NormalCases)
    AbnormalCases = np.array(dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"])
    K2 = len(AbnormalCases)
    K = K1 + K2

    # calculate lesions perCase
    x = pd.Series(dfTruth["CaseID"])
    x = x.isin(AbnormalCases)
    x = pd.Series(dfTruth["CaseID"][x])
    x = x.value_counts(sort=False)
    perCase = x.sort_index()
    # next line indexes using abnormal cases, 0 to K2-1
    # Fixes indexing of perCase array
    # perCase = pd.Series(list(perCase))
    # did not need the above after all
    perCase = np.array(perCase)

    maxLL = max(perCase)
    if (maxLL > 1) & (DataType != "FROC"):
        sys.exit("Only FROC data can have more than one lesion per case")
    relWeights = np.array([1/maxLL] * maxLL)

# =============================================================================
# Load the NL sheet
# Extract the columns
# =============================================================================
    ws = wb['NL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfNL = pd.DataFrame(data, columns=columnNames)

# =============================================================================
# Load the LL sheet
# Exract the columns
# =============================================================================
    ws = wb['LL']
    data = ws.values
    columnNames = next(data)[0:]
    # Extract the data minus the column names
    dfLL = pd.DataFrame(data, columns=columnNames)

    # with FROC data not all modalities may appear in NL and LL sheets
    modalities = (dfLL["ModalityID"].append(dfNL["ModalityID"])).unique()
    I = len(modalities)
    modalityID = [str(x) for x in list(range(I))]
    modalityID = np.array(modalityID)
    readers = (dfLL["ReaderID"].append(dfNL["ReaderID"])).unique()
    J = len(readers)
    readerID = [str(x) for x in list(range(J))]
    readerID = np.array(readerID)
    # lesions = np.unique(dfTruth["LesionID"])[1:]
    maxNL = dfNL.groupby(['ReaderID',
                          'ModalityID',
                          'CaseID']).transform(len).max()[0]

    if (maxNL > 1) & (DataType != "FROC"):
        sys.exit("Only FROC data can have more than one NL per case")

    NL = np.full((I, J, K, maxNL), -np.inf)
    indxNl = 0
    while indxNl < len(dfNL["ModalityID"]):
    #for indxNl in range(len(dfNL["ModalityID"])):
        i = (modalities == dfNL["ModalityID"][indxNl])
        j = (readers == dfNL["ReaderID"][indxNl])
        c = (AllCases == dfNL["CaseID"][indxNl])

        matchCount = ((dfNL["CaseID"] == dfNL["CaseID"][indxNl]) &
                      (dfNL["ModalityID"] == dfNL["ModalityID"][indxNl]) &
                      (dfNL["ReaderID"] == dfNL["ReaderID"][indxNl]))
        for l in range(sum(matchCount)):
            if NL[i, j, c, l] == -np.inf:
                NL[i, j, c, l] = dfNL["NLRating"][indxNl + l]
            else:
                print("i: %2d, j: %2d, c: %4d" % (i, j, c))
                sys.exit("Overwriting NL values!")
        indxNl += sum(matchCount)

    lesions = list(range(1, maxLL+1))
    LL = np.full((I, J, K2, maxLL), -np.inf)
    for indxLl in range(len(dfLL["ModalityID"])):
        i = (modalities == dfLL["ModalityID"][indxLl])
        j = (readers == dfLL["ReaderID"][indxLl])
        c = (AbnormalCases == dfLL["CaseID"][indxLl])
        l = (lesions == dfLL["LesionID"][indxLl])
        if LL[i, j, c, l] == -np.inf:
            LL[i, j, c, l] = dfLL["LLRating"][indxLl]
        else:
            print("i: %2d, j: %2d, c: %4d" % (i, j, c))
            sys.exit("Overwriting LL values!")
        
# =============================================================================
# check that no abnormal case has more marked lesions that total number of les
# =============================================================================
    for i in range(I):
        for j in range(J):
            for c in range(K2):
                if sum(LL[i,j,c,:] != -np.inf) > perCase[c]:
                    print("i: %2d, j: %2d, c: %4d" % (i, j, c))
                    sys.exit("number of LLs greater than number of lesions")

# =============================================================================
# Return a dataset object
# =============================================================================
    ds = [NL, LL, perCase, relWeights, DataType, modalityID, readerID]
    return(ds)


# FileName = "extdata/toyFiles/FROC/frocCr.xlsx"
# FileName = "extdata/JT.xlsx"
# ds = DfReadDataFile(FileName)


def DfExtractDataset(ds, trts = None, rdrs = None):
    """
    Extract a subset of treatments/readers from a larger dataset

    Parameters
    ----------
    ds : list
        The original dataset from which the subset is to be extracted

    trts : list 
        The indices of the treatments to extract; if missing then all 
        treatments are extracted

    trts : list 
        The indices of the readers to extract; 
        if missing then all readers are extracted

    Returns
    -------
    A dataset containing only the specified treatments and readers extracted from the original dataset

    """
    
    if (trts == None) & (rdrs == None):
        return ds
    
    NL = ds[0]
    LL = ds[1]
    I = len(NL[:, 0, 0, 0])
    J = len(NL[0, :, 0, 0])

    perCase = ds[2]
    relWeights = ds[3]
    DataType = ds[4]
    
    if trts == None:
        trts = list(range(I))
    if rdrs == None:
        rdrs = list(range(J))

    # e denotes extracted values
# =============================================================================
#   TODO Tnere is a risk of losing dimensions; need more careful checking with
#   other datasets
# =============================================================================
    # this has to be done in two steps !!why?!! !!dpc!!
    NLe = NL[trts, :, :, :]
    NLe = NLe[:, rdrs, :, :]
    LLe = LL[trts, :, :, :]
    LLe = LLe[:, rdrs, :, :]
    # TODO test with dataset where one reader produces lots more NLs than others
    # Extracting all but this reader will reduce maxNL and may break this code
# =============================================================================
# Next 2 lines breaks C-contiguous property of NLe:
# NLe.flags['C_CONTIGUOUS'] -> False 
# dont really need them; they appear unnecessary
# first line determines length of the 4th dimensiom
# and the next operation is picking all elements of the 4th dimension   
    ## maxNLe = len(NLe[0, 0, 0, :])
    ## NLe = NLe[:,:,:,list(range(maxNLe))]
# =============================================================================
    # dont need above type code as maxLL is fixed by Truth sheet and is 
    # independent of treatment or reader
    # maxLL = len(LLe[0, 0, 0, :])

    modalityIDe = []
    for i in range(len(trts)):
        modalityIDe.append(trts[i])
    modalityIDe = [str(x) for x in modalityIDe]
    
    readerIDe = []
    for i in range(len(rdrs)):
        readerIDe.append(rdrs[i])
    readerIDe = [str(x) for x in readerIDe]

    pass

    dse = [NLe, LLe, perCase, relWeights, DataType, modalityIDe, readerIDe]
    return(dse)


def DfFroc2Roc(ds):
    """
    Convert an FROC dataset to an ROC dataset

    Parameters
    ----------
    ds : list
        The original FROC dataset


    Returns
    -------
    A highest-rating inferred-ROC dataset

    """
    NL = ds[0]
    LL = ds[1]
    K = len(NL[0, 0, :, 0])
    K2 = len(LL[0, 0, :, 0])
    K1 = K - K2

    ds1 = copy.deepcopy(ds)
    nl = np.array(ds1[0])
    fp = np.max(nl, axis=3, keepdims=True)
    nl1 = fp[:, :, K1:, :]
    ll = np.array(ds1[1])
    ll1 = np.max(ll, axis=3, keepdims=True)
    tp = np.maximum(nl1, ll1)
    fp[fp == -np.inf] = 0
    tp[tp == -np.inf] = 0
    ds1[0] = fp
    ds1[1] = tp
    ds1[2] = pd.Series([1] * K2)
    ds1[3] = [1]
    ds1[4] = "ROC"
    pass
    return ds1


def DfRatings2Dataset (NL, LL, InputIsCountsTable = False, **kwargs):
    """
    Convert ratings arrays to a dataset; currently limited to single modality
    single reader dataset

    Parameters
    ----------
    NL : array
        Non-lesion localizations array (or FP array for ROC data)

    LL : array
        Lesion localizations array (or TP array for ROC data)
    
    InputIsCountsTable: boolean 
        If True, the NL and LL arrays are rating-counts tables, with common 
        lengths equal to the number of ratings R, if False, the default, 
        these are arrays of lengths K1, the number of non-diseased cases, 
        and K2, the number of diseased cases, respectively.
        
    ...: optional parameters
        Other elements of the dataset that may, depending on the context, 
        need to be specified. perCase must be specified if an FROC dataset is 
        to be returned. It is a K2-length array specifying the numbers of 
        lesions in each diseased case in the dataset.

    Returns
    -------
    ds: A dataset

    """

    if NL.ndim != LL.ndim:
        sys.exit("number of dimensions of NL and LL arrays must be equal")

    if NL.ndim > 1:
        sys.exit("unimplemented code: limited to single modality and single reader")
        
    if len(kwargs) == 0:
        if InputIsCountsTable:
            R = len(NL)
            if (len(LL) != R):
                sys.exit("Lengths of rating-counts arrays are unequal")
            NL = [j+1 for j in range(R) for i in range(NL[j])]
            LL = [j+1 for j in range(R) for i in range(LL[j])]
            I = 1
            J = 1
            K1 = len(NL)
            K2 = len(LL)
            perCase = np.ones(K2)
            NL1 = np.full((I, J, K1+K2, 1), -np.inf)
            NL1[0,0,0:K1,0] = NL
            LL1 = np.full((I, J, K2, 1), -np.inf)
            LL1[0,0,:,0] = LL
            relWeights = 1
            modalityID = "0"
            readerID = "0"
            DataType = "ROC" 
            ds = [NL1, LL1, perCase, relWeights, DataType, modalityID, readerID]
            return ds
    else:
        for key in kwargs:
            [names, values] = (key, kwargs[key])
        if names != "perCase":
            sys.exit("missing 'perCase' argument")
        else: perCase = values
        
        if NL.ndim == 1: #single modality and single reader
            I = 1
            J = 1
            K1 = len(NL)
            K2 = len(LL)
            if len(perCase) != K2:
                sys.exit("length of perCase not equal to len(LL)")
            NL1 = np.full((I, J, K1+K2, 1), -np.inf)
            NL1[0,0,0:K1,0] = NL
            LL1 = np.full((I, J, K2, 1), -np.inf)
            LL1[0,0,:,0] = LL
            DataType = "FROC"
            relWeights = 1
            modalityID = "0"
            readerID = "0"
            ds = [NL1, LL1, perCase, relWeights, DataType, modalityID, readerID]
            return ds
            
